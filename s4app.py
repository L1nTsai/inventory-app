import tkinter as tk
import os
from tkinter import ttk, messagebox, simpledialog
from openpyxl import Workbook, load_workbook
from datetime import datetime
from PIL import Image, ImageTk



class MainMenuApp:
    def __init__(self, root):
        self.root = root
        self.root.title("S4 Inventory Tool")

        self.create_widgets()


    def create_widgets(self):
        tk.Label(self.root, text="Main Menu", font=('Arial', 36)).pack(pady=20)
        tk.Button(self.root, text='Inventory', cursor='hand1',font=('Arial', 24), command=self.run_inventory_app, width=15, height=2).pack(pady=10)
        tk.Button(self.root, text='Checkout', cursor='hand1',font=('Arial', 24), command=self.run_checkout_app, width=15, height=2).pack(pady=10)
        tk.Button(self.root, text='History', cursor='hand1',font=('Arial', 24), command=self.run_history_app, width=15, height=2).pack(pady=10)
        tk.Button(self.root, text='Files', cursor='hand1',font=('Arial', 24), command=self.open_location, width=15, height=2).pack(pady=10)
        tk.Button(self.root, text='Getting Started', cursor='hand1',font=('Arial', 24), command=self.run_instructions, width=15, height=2).pack(pady=10)

    def run_instructions(self):
        self.root.withdraw()
        instructions_root = tk.Toplevel(self.root)
        instructions_page = InstructionsPage(instructions_root, self.show_main_menu)

    def run_inventory_app(self):
        self.root.withdraw()  # Hide the main menu
        inventory_root = tk.Toplevel(self.root)
        inventory_app = InventoryApp(inventory_root, self.show_main_menu)

    def run_checkout_app(self):
        self.root.withdraw()  # Hide the main menu
        checkout_root = tk.Toplevel(self.root)
        checkout_app = CheckoutApp(checkout_root, self.show_main_menu)

    def run_history_app(self):
        self.root.withdraw()  # Hide the main menu
        history_root = tk.Toplevel(self.root)
        history_app = HistoryApp(history_root, self.show_main_menu)
     
    def open_location(self):
        self.root.withdraw()  # Hide the main menu
        files_root = tk.Toplevel(self.root)
        files_app = FileTab(files_root, self.show_main_menu)

    def show_main_menu(self):
        self.root.deiconify()  # Show the main menu when called

class InstructionsPage:
    def __init__(self, root, back_callback):
        self.root = root
        self.root.title("Instructions")

        self.back_callback = back_callback

        self.create_widgets()
    
    def create_widgets(self):
        tk.Label(self.root, text="If this is your first time running this app, \n please click 'inventory' and 'history' to create all the necessary files. \n App made by Lin Tsai", font=('Arial', 36)).pack(pady=20)
        tk.Button(self.root, text='Back', font=('Arial', 24), command=self.back_to_menu, width=15, height=2).pack(pady=10)

    
    def back_to_menu(self):
        self.root.destroy()  # Destroy the current file tab window
        self.back_callback()  # Show the main menu



class FileTab:
    def __init__(self, root, back_callback):
        self.root = root
        self.root.title("File Tab")

        self.back_callback = back_callback

        self.create_widgets()

    def create_widgets(self):
        tk.Label(self.root, text="File Tab", font=('Arial', 36)).pack(pady=20)

        # Combobox for selecting the operating system
        self.os_var = tk.StringVar()
        os_options = ['Windows', 'macOS/Linux']
        tk.Label(self.root, text='Select Operating System:').pack()
        os_combobox = ttk.Combobox(self.root, textvariable=self.os_var, values=os_options)
        os_combobox.pack(pady=10)
        os_combobox.set(os_options[0])  # Set the default value

        # Button to open the location based on the selected operating system
        tk.Button(self.root, text='Open Location', cursor='hand1',font=('Arial', 24), command=self.open_location, width=15, height=2).pack(pady=10)
        
        # Back button
        tk.Button(self.root, text='Back', cursor='hand1',font=('Arial', 24), command=self.back_to_menu, width=15, height=2).pack(pady=10)

    def open_location(self):
        selected_os = self.os_var.get()

        script_dir = os.path.dirname(os.path.abspath(__file__))
        if selected_os == 'Windows':
            os.system(f'explorer {script_dir}')
        elif selected_os == 'macOS/Linux':
            os.system(f'open {script_dir}')  # For macOS
            # Alternatively, you can use the following for Linux:
            # os.system(f'xdg-open {script_dir}')

    def back_to_menu(self):
        self.root.destroy()  # Destroy the current file tab window
        self.back_callback()  # Show the main menu


class InventoryApp:
    def __init__(self, root, back_callback):
        self.root = root
        self.root.title("Inventory App")

        self.back_callback = back_callback

        self.create_widgets()
        self.load_inventory()

    def create_widgets(self):
        # Entry for item name, quantity, and item id
        self.item_entry = tk.Entry(self.root)
        self.quantity_entry = tk.Entry(self.root)
        self.id_entry = tk.Entry(self.root)

        tk.Label(self.root, text='Item Name:').grid(row=0, column=0, sticky=tk.E)
        self.item_entry.grid(row=0, column=1)

        tk.Label(self.root, text='Item ID:').grid(row=0, column=2, sticky=tk.E)
        self.id_entry.grid(row=0, column=3)

        tk.Label(self.root, text='Quantity:').grid(row=1, column=0, sticky=tk.E)
        self.quantity_entry.grid(row=1, column=1)

        # Buttons for adding, updating, removing items, clearing spreadsheet, and going back
        tk.Button(self.root, text='Add Item', cursor='hand1',command=self.add_item, width=15, height=2).grid(row=2, column=1, pady=5)
        tk.Button(self.root, text='Remove Selected Item', cursor='hand1',command=self.remove_selected_item, width=15, height=2).grid(row=2, column=3, pady=5)
        tk.Button(self.root, text='Update Selected Item', cursor='hand1',command=self.update_selected_item, width=15, height=2).grid(row=3, column=1, pady=5)
        tk.Button(self.root, text='Clear Spreadsheet', cursor='hand1',command=self.clear_spreadsheet, width=15, height=2).grid(row=3, column=3, pady=5)
        tk.Button(self.root, text='Back', cursor='hand1',command=self.back_to_menu, width=15, height=2).grid(row=4, column=1, pady=10)

        # Treeview to display inventory
        self.tree = ttk.Treeview(self.root, columns=('', 'ID', 'Item Name', 'Quantity'))
        self.tree.heading('#0', text='')
        self.tree.heading('#1', text='ID')
        self.tree.heading('#2', text='Item Name')
        self.tree.heading('#3', text='Quantity')
        self.tree.column('#0', width=0)
        self.tree.grid(row=0, column=4, rowspan=4, padx=10, pady=10)
        self.tree.bind('<ButtonRelease-1>', self.on_tree_select)  # Bind the Treeview selection event

    def load_inventory(self):
        try:
            self.workbook = load_workbook('inventory.xlsx')
            self.sheet = self.workbook.active
            self.update_treeview()
        except FileNotFoundError:
            # Create a new workbook if the file doesn't exist
            self.workbook = Workbook()
            self.sheet = self.workbook.active
            self.sheet.append(['ID', 'Item Name', 'Quantity'])
            self.workbook.save('inventory.xlsx')

    def update_treeview(self):
        # Clear existing items in the treeview
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Populate treeview with data from the Excel file
        for row in self.sheet.iter_rows(min_row=2, values_only=True):
            self.tree.insert('', 'end', values=row)

    def add_item(self):
        item = self.item_entry.get()
        quantity = self.quantity_entry.get()
        item_id = self.id_entry.get()

        if item and quantity and item_id:
            try:
                row_id = int(item_id)
            except ValueError:
                messagebox.showerror('Error', 'Item ID must be an integer.')
                return

            # Check if the ID already exists in the sheet
            for existing_row in self.sheet.iter_rows(min_row=2, values_only=True):
                if existing_row[0] == row_id:
                    # Update the existing row
                    self.sheet.cell(row=row_id + 1, column=2, value=item)
                    self.sheet.cell(row=row_id + 1, column=3, value=int(quantity))
                    self.workbook.save('inventory.xlsx')
                    self.update_treeview()
                    messagebox.showinfo('Info', f'Item ID {row_id} updated successfully.')
                    return

            # If the ID doesn't exist, add a new row
            self.sheet.append([row_id, item, int(quantity)])
            self.workbook.save('inventory.xlsx')
            self.update_treeview()
            messagebox.showinfo('Info', 'Item added successfully.')
        else:
            messagebox.showwarning('Warning', 'Please enter values for item name, quantity, and item ID.')

        # Clear entry fields
        self.item_entry.delete(0, 'end')
        self.quantity_entry.delete(0, 'end')
        self.id_entry.delete(0, 'end')

    def update_selected_item(self):
        selected_item = self.tree.selection()
        if selected_item:
            item_id = self.tree.item(selected_item)['values'][0]
            new_item = self.item_entry.get()
            new_quantity = self.quantity_entry.get()

            if new_item and new_quantity:
                for index, row in enumerate(self.sheet.iter_rows(min_row=2, max_row=self.sheet.max_row, values_only=True), start=2):
                    if row[0] == item_id:
                        # Update the existing row
                        self.sheet.cell(row=index, column=2, value=new_item)
                        self.sheet.cell(row=index, column=3, value=int(new_quantity))
                        self.workbook.save('inventory.xlsx')
                        self.update_treeview()
                        messagebox.showinfo('Info', f'Item ID {item_id} updated successfully.')
                        return
                else:
                    messagebox.showwarning('Warning', 'Selected item not found.')
            else:
                messagebox.showwarning('Warning', 'Please enter both item name and quantity.')
        else:
            messagebox.showwarning('Warning', 'Please select an item from the inventory.')

    def remove_selected_item(self):
        selected_item = self.tree.selection()
        if selected_item:
            item_id = self.tree.item(selected_item)['values'][0]
            for index, row in enumerate(self.sheet.iter_rows(min_row=2, max_row=self.sheet.max_row, values_only=True), start=2):
                if row[0] == item_id:
                    self.sheet.delete_rows(index)
                    self.workbook.save('inventory.xlsx')
                    self.update_treeview()
                    messagebox.showinfo('Info', f'Item ID {item_id} removed successfully.')
                    return
        else:
            messagebox.showwarning('Warning', 'Please select an item from the inventory.')

    def clear_spreadsheet(self):
        confirmation = messagebox.askyesno('Confirmation', 'Are you sure you want to clear the spreadsheet? This action cannot be undone.')
        if confirmation:
            # Clear all rows except the first row (headers)
            for row in range(self.sheet.max_row, 1, -1):
                self.sheet.delete_rows(row)
            self.workbook.save('inventory.xlsx')
            self.update_treeview()
            messagebox.showinfo('Info', 'Spreadsheet cleared successfully.')
        else:
            messagebox.showinfo('Info', 'Spreadsheet was not cleared.')

    def on_tree_select(self, event):
        selected_item = self.tree.selection()
        if selected_item:
            item_id, item, quantity = self.tree.item(selected_item)['values']
            self.id_entry.delete(0, 'end')
            self.item_entry.delete(0, 'end')
            self.quantity_entry.delete(0, 'end')

            self.id_entry.insert(0, item_id)
            self.item_entry.insert(0, item)
            self.quantity_entry.insert(0, quantity)

    def back_to_menu(self):
        self.root.destroy()  # Destroy the current inventory window
        self.back_callback()  # Show the main menu

class CheckoutApp:
    def __init__(self, root, back_callback):
        self.root = root
        self.root.title("Checkout App")

        self.back_callback = back_callback
        self.cart = []

        # Load the history workbook
        try:
            self.history_workbook = load_workbook('history.xlsx')
            self.history_sheet = self.history_workbook.active
        except FileNotFoundError:
            messagebox.showerror('Error', 'History file not found. Please run the History App first.')
            self.root.destroy()

        self.create_widgets()
        self.load_inventory()

    def create_widgets(self):
        # Entry for student ID, item ID, quantity, and Listbox to display cart items
        self.student_id_entry = tk.Entry(self.root)
        self.item_id_entry = tk.Entry(self.root)
        self.quantity_entry = tk.Entry(self.root)

        tk.Label(self.root, text='Student ID:').grid(row=0, column=0, sticky=tk.E)
        self.student_id_entry.grid(row=0, column=1)

        tk.Label(self.root, text='Item ID:').grid(row=1, column=0, sticky=tk.E)
        self.item_id_entry.grid(row=1, column=1)

        tk.Label(self.root, text='Quantity:').grid(row=2, column=0, sticky=tk.E)
        self.quantity_entry.grid(row=2, column=1)

        tk.Button(self.root, text='Add to Cart', cursor='hand1',command=self.add_to_cart, width=15, height=2).grid(row=4, column=0, pady=5)
        tk.Button(self.root, text='Checkout', cursor='hand1',command=self.checkout, width=15, height=2).grid(row=4, column=1, pady=5)
        tk.Button(self.root, text='Back', cursor='hand1',command=self.back_to_menu, width=15, height=2).grid(row=5, column=1, pady=10)

        # Listbox to display items in the cart
        self.cart_listbox = tk.Listbox(self.root, selectmode=tk.SINGLE, width=40, height=10)
        self.cart_listbox.grid(row=3, column=0, padx=10, pady=10, columnspan=2)


    def load_inventory(self):
        try:
            self.workbook = load_workbook('inventory.xlsx')
            self.sheet = self.workbook.active
        except FileNotFoundError:
            messagebox.showerror('Error', 'Inventory file not found. Please run the Inventory App first.')
            self.root.destroy()

    def add_to_cart(self):
        item_id_str = self.item_id_entry.get()
        quantity_str = self.quantity_entry.get()

        if item_id_str and quantity_str:
            try:
                item_id = int(item_id_str)
                quantity = int(quantity_str)
            except ValueError:
                messagebox.showerror('Error', 'Item ID and Quantity must be integers.')
                return

            # Check if the item exists in the inventory
            for row in self.sheet.iter_rows(min_row=2, values_only=True):
                if row[0] == item_id:
                    if quantity > 0 and quantity <= row[2]:
                        self.cart.append((item_id, row[1], quantity))
                        self.cart_listbox.insert(tk.END, f'{row[1]} (ID: {item_id}) - Quantity: {quantity}')
                        messagebox.showinfo('Info', f'Item ID {item_id} added to cart.')
                        return
                    else:
                        messagebox.showwarning('Warning', f'Invalid quantity for item ID {item_id}.')
                        return

            messagebox.showwarning('Warning', f'Item ID {item_id} not found in inventory.')
        else:
            messagebox.showwarning('Warning', 'Please enter both Item ID and Quantity.')

    def checkout(self):
        student_id_str = self.student_id_entry.get()
        if student_id_str and self.cart:
            try:
                student_id = int(student_id_str)
            except ValueError:
                messagebox.showerror('Error', 'Student ID must be an integer.')
                return

            # Update the inventory, record the transaction in history, and clear the cart
            for cart_item in self.cart:
                item_id = cart_item[0]
                quantity_in_cart = cart_item[2]

                # Update inventory
                for index, inventory_row in enumerate(self.sheet.iter_rows(min_row=2, max_row=self.sheet.max_row, values_only=True), start=2):
                    if inventory_row[0] == item_id:
                        inventory_quantity = inventory_row[2]
                        if quantity_in_cart <= inventory_quantity:
                            self.sheet.cell(row=index, column=3, value=inventory_quantity - quantity_in_cart)
                        else:
                            messagebox.showwarning('Warning', f'Not enough quantity in inventory for item ID {item_id}.')
                        break

                # Record transaction in history
                current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                self.history_sheet.append([current_time, student_id, item_id, cart_item[1], quantity_in_cart])

            self.workbook.save('inventory.xlsx')
            self.history_workbook.save('history.xlsx')
            self.cart = []  # Clear the cart
            self.cart_listbox.delete(0, tk.END)  # Clear the listbox
            self.student_id_entry.delete(0, tk.END)  # Clear the student ID entry
            messagebox.showinfo('Info', 'Checkout successful.')
        else:
            messagebox.showwarning('Warning', 'Please enter a student ID and add items to the cart.') 

    def back_to_menu(self):
        self.root.destroy()  # Destroy the current checkout window
        self.back_callback()  # Show the main menu

class HistoryApp:
    def __init__(self, root, back_callback):
        self.root = root
        self.root.title("History App")

        self.back_callback = back_callback

        # Create history spreadsheet if not exists
        try:
            self.workbook = load_workbook('history.xlsx')
        except FileNotFoundError:
            self.workbook = Workbook()
            self.workbook.save('history.xlsx')

        self.history_sheet = self.workbook.active
        self.create_widgets()
        self.update_treeview()

    def create_widgets(self):
        # Frame for entry and buttons
        entry_frame = tk.Frame(self.root)
        entry_frame.grid(row=0, column=0, padx=10, pady=10)

        # Entry for searching by student ID
        self.search_entry = tk.Entry(entry_frame)
        self.search_entry.grid(row=0, column=0, padx=10, pady=10)

        # Button to search
        tk.Button(entry_frame, text='Search', cursor='hand1',command=self.search_history, width=10, height=1).grid(row=0, column=1, padx=10, pady=10)

        # Button to clear history
        tk.Button(entry_frame, text='Clear History', cursor='hand1',command=self.clear_history, width=15, height=2).grid(row=1, column=0, pady=10)

        # Button to back to menu
        tk.Button(entry_frame, text='Back to Menu', cursor='hand1',command=self.back_to_menu, width=15, height=2).grid(row=1, column=1, pady=10)

        # Treeview to display history
        self.tree = ttk.Treeview(self.root, columns=('Time', 'Student ID', 'Item ID', 'Item Name', 'Quantity'))
        self.tree.heading('#0', text='')
        self.tree.heading('#1', text='Time')
        self.tree.heading('#2', text='Student ID')
        self.tree.heading('#3', text='Item ID')
        self.tree.heading('#4', text='Item Name')
        self.tree.heading('#5', text='Quantity')
        self.tree.column('#0', width=0)
        self.tree.grid(row=0, column=1, padx=10, pady=10, rowspan=3)

    def update_treeview(self):
        # Clear existing items in the treeview
        for item in self.tree.get_children():
            self.tree.delete(item)

        # Populate treeview with data from the history spreadsheet
        for row in self.history_sheet.iter_rows(min_row=2, values_only=True):
            self.tree.insert('', 'end', values=row)

    def search_history(self):
        search_value = self.search_entry.get().strip()
        if search_value:
            # Clear existing items in the treeview
            for item in self.tree.get_children():
                self.tree.delete(item)

            # Populate treeview with filtered data from the history spreadsheet
            for row in self.history_sheet.iter_rows(min_row=2, values_only=True):
                if str(row[1]) == search_value:
                    self.tree.insert('', 'end', values=row)
        else:
            # If the search value is empty, show all rows
            self.update_treeview()

    def clear_history(self):
        passcode = simpledialog.askstring('Passcode', 'Enter the passcode:', show='*')
        if passcode == '123':
            # Clear all rows except the first row (headers)
            for row in range(self.history_sheet.max_row, 1, -1):
                self.history_sheet.delete_rows(row)
            self.workbook.save('history.xlsx')
            self.update_treeview()
            messagebox.showinfo('Info', 'History cleared successfully.')
        else:
            messagebox.showinfo('Info', 'Incorrect passcode. History was not cleared.')

    def back_to_menu(self):
        self.root.destroy()  # Destroy the current history window
        self.back_callback()  # Show the main menu



if __name__ == "__main__":
    root = tk.Tk()
    root.title("Main Menu")
    root.geometry("800x600")  # Set the desired size
    app = MainMenuApp(root)
    root.iconbitmap('omi.ico')
    root.mainloop()
    

